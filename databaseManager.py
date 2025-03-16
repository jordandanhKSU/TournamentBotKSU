import os
import aiosqlite
from openpyxl import load_workbook
from dotenv import load_dotenv, find_dotenv
import discord
import aiohttp
import asyncio

load_dotenv(find_dotenv())
DB_PATH = os.getenv('DB_PATH')
SPREADSHEET_PATH = os.path.abspath(os.getenv('SPREADSHEET_PATH'))

async def get_db_connection():
    return await aiosqlite.connect(DB_PATH)

async def initialize_database():
    async with aiosqlite.connect(DB_PATH) as conn:
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS "PlayerStats" (
                "DiscordID" TEXT NOT NULL UNIQUE,
                "DiscordUsername" TEXT NOT NULL,
                "PlayerRiotID" TEXT UNIQUE,
                "Participation" NUMERIC DEFAULT 0,
                "Wins" INTEGER DEFAULT 0,
                "MVPs" INTEGER DEFAULT 0,
                "ToxicityPoints" NUMERIC DEFAULT 0,
                "GamesPlayed" INTEGER DEFAULT 0,
                "WinRate" REAL,
                "TotalPoints" NUMERIC DEFAULT 0,
                "PlayerTier" INTEGER DEFAULT 0,
                "PlayerRank" TEXT DEFAULT 'UNRANKED',
                "RolePreference" TEXT DEFAULT '55555',
                PRIMARY KEY("DiscordID")
            )
        ''')
        await conn.commit()

def update_excel(discord_id, player_data):
    try:
        # Load the workbook and get the correct sheet
        workbook = load_workbook(SPREADSHEET_PATH)
        sheet_name = 'PlayerStats'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f'Sheet {sheet_name} does not exist in the workbook')

        # Check if the player already exists in the sheet
        found = False
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
            if str(row[0].value) == discord_id:  # Check if Discord ID matches
                # Update only if there's a difference
                for idx, key in enumerate(player_data.keys()):
                    if row[idx].value != player_data[key]:
                        row[idx].value = player_data[key]
                        found = True
                break

        # If player not found, add a new row
        if not found:
            # Find the first truly empty row, ignoring formatting
            empty_row_idx = sheet.max_row + 1
            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if all(cell.value is None for cell in row):
                    empty_row_idx = i
                    break

            # Insert the new data into the empty row
            for idx, key in enumerate(player_data.keys(), start=1):
                sheet.cell(row=empty_row_idx, column=idx).value = player_data[key]

        # Save the workbook after updates
        workbook.save(SPREADSHEET_PATH)
        print(f"Spreadsheet '{SPREADSHEET_PATH}' has been updated successfully.")

    except Exception as e:
        print(f"Error updating Excel file: {e}")

async def check_winners_in_db(winners):
    not_found_users = []
    async with aiosqlite.connect(DB_PATH) as conn:
        for winner in winners:
            # Use winner.discord_id instead of winner.id
            async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(winner.discord_id),)) as cursor:
                result = await cursor.fetchone()

            if not result:
                # Add to not found list
                not_found_users.append(winner)


async def update_wins(winners):
    async with aiosqlite.connect(DB_PATH) as conn:
        for winner in winners:
            # Use winner.discord_id instead of winner.id
            async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(winner.discord_id),)) as cursor:
                result = await cursor.fetchone()
                if result:
                    wins, games_played = result
                    await conn.execute(
                        "UPDATE PlayerStats SET Wins = ?, GamesPlayed = ? WHERE DiscordID = ?",
                        (wins + 1, games_played + 1, str(winner.discord_id))
                    )
                    await update_win_rate(conn, str(winner.discord_id))
        await conn.commit()



# code to calculate and update winrate in database
async def update_win_rate(conn, discord_id):
    async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (discord_id,)) as cursor:
        result = await cursor.fetchone()
    if result:
        wins, games_played = result
        win_rate = (wins / games_played) * 100 if games_played > 0 else 0
        await conn.execute("UPDATE PlayerStats SET WinRate = ? WHERE DiscordID = ?", (win_rate, discord_id))

# discord roles need to be implemented here
# maybe adding functionality for multiple users 
async def update_points(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()
        
        if result is None:
            return f"Failed to update: {member} (user not found in database)."
        
        await conn.execute("UPDATE PlayerStats SET Participation = Participation + 1, TotalPoints = TotalPoints + 1 WHERE DiscordID = ?", (member,))
        await conn.commit()
        return f"Added 1 participation point to {member}."

async def update_username(player):
    try:
        current_username = player.display_name
        async with aiosqlite.connect(DB_PATH) as conn:
            # Fetch the player's current data from the database
            async with conn.execute(
                "SELECT DiscordUsername FROM PlayerStats WHERE DiscordID = ?",
                (str(player.id),)
            ) as cursor:
                player_stats = await cursor.fetchone()

            # If player exists in the database and the username is outdated, update it
            if player_stats:
                stored_username = player_stats[0]
                if stored_username != current_username:
                    await conn.execute(
                        "UPDATE PlayerStats SET DiscordUsername = ? WHERE DiscordID = ?",
                        (current_username, str(player.id))
                    )
                    await conn.commit()
                    print(f"Updated username for {player.id} to '{current_username}'.")
            else:
                # Insert new player if they don't exist
                await conn.execute(
                    """INSERT INTO PlayerStats (DiscordID, DiscordUsername)
                       VALUES (?, ?)""",
                    (str(player.id), current_username)
                )
                await conn.commit()
                print(f"Added new player {player.id} with username '{current_username}'.")
    except Exception as e:
        print(f"Error updating username: {e}")

# member was interaction: discord.Interaction
async def link(member, riot_id: str):
    """
    Link a Discord user's account with their Riot ID.
    
    Args:
        member: The Discord member object
        riot_id: The Riot ID in format 'username#tagline'
    
    Returns:
        A message string indicating the result of the operation
    """
    # Riot ID is in the format 'username#tagline', e.g., 'jacstogs#1234'
    if '#' not in riot_id:
        return "Invalid Riot ID format. Must include a '#' character (e.g., 'username#tagline')."

    # Split the Riot ID into name and tagline
    summoner_name, tagline = riot_id.split('#', 1)
    summoner_name = summoner_name.strip()
    tagline = tagline.strip()
    
    # Check if either part is empty
    if not summoner_name or not tagline:
        return "Invalid Riot ID format. Both username and tagline must be provided."

    # Check if API key exists
    api_key = os.getenv("RIOT_API_KEY")
    if not api_key:
        return "ERROR: Riot API key not found. Please add a valid API key to your environment variables."
    
    # Verify that the Riot ID exists using the Riot API
    headers = {"X-Riot-Token": api_key}
    url = f"https://americas.api.riotgames.com/riot/account/v1/accounts/by-riot-id/{summoner_name}/{tagline}"
    verified = False  # Default to not verified

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    # Riot ID exists
                    data = await response.json()
                    verified = True
                else:
                    # Failed to verify Riot ID
                    error_msg = await response.text()
                    if response.status == 404:
                        return f"The Riot ID '{riot_id}' could not be found. Please double-check and try again."
                    elif response.status == 400:
                        return f"ERROR: Bad Request. This is likely due to an invalid or expired API key format."
                    elif response.status == 401:
                        return f"ERROR: Invalid Riot API key. Please verify your API key or obtain a new one from the Riot Developer Portal."
                    elif response.status == 403:
                        return f"ERROR: Riot API key has expired or doesn't have sufficient permissions. Please obtain a new key from the Riot Developer Portal."
                    else:
                        return f"ERROR: API Error (status {response.status}). This may be due to API key issues or service problems."
                    
    except aiohttp.ClientError as e:
        return f"Network error connecting to Riot API: {e}. Please check your internet connection and try again."
    except Exception as e:
        return f"Unexpected error during API validation: {e}. Please try again later."

    # Only proceed if verification passed
    if verified:
        async with aiosqlite.connect(DB_PATH) as conn:
            try:
                # Use member's id for database operations
                async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
                    result = await cursor.fetchone()

                if result:
                    # Update the existing record with the new Riot ID
                    await conn.execute(
                        "UPDATE PlayerStats SET PlayerRiotID = ? WHERE DiscordID = ?",
                        (riot_id, str(member.id))
                    )
                else:
                    # Insert a new record if the user doesn't exist in the database
                    await conn.execute(
                        "INSERT INTO PlayerStats (DiscordID, DiscordUsername, PlayerRiotID) VALUES (?, ?, ?)",
                        (str(member.id), member.display_name, riot_id)
                    )

                await conn.commit()
                return f"Your Riot ID '{riot_id}' has been successfully linked to your Discord account."

            except aiosqlite.IntegrityError as e:
                # Handle UNIQUE constraint violation (i.e., Riot ID already linked)
                if 'UNIQUE constraint failed: PlayerStats.PlayerRiotID' in str(e):
                    # Riot ID is already linked to another user
                    async with conn.execute("""
                        SELECT DiscordID, DiscordUsername FROM PlayerStats WHERE PlayerRiotID = ?
                    """, (riot_id,)) as cursor:
                        existing_user_data = await cursor.fetchone()

                    if existing_user_data:
                        existing_user_id, existing_username = existing_user_data
                        return f"Error: This Riot ID is already linked to another Discord user: <@{existing_user_id}>. If this is a mistake, please contact an administrator."
                else:
                    return f"Database error: {e}"
            except Exception as e:
                return f"An unexpected error occurred while updating the database: {e}"
    # This else clause is not needed since all code paths in the if block return a value

async def update_toxicity_by_id(discord_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Search for the user in the PlayerStats table using the given discord_id
        async with conn.execute("SELECT ToxicityPoints FROM PlayerStats WHERE DiscordID = ?", (discord_id,)) as cursor:
            result = await cursor.fetchone()

        if result:
            toxicity_points = result[0]
            # Increment the toxicity points and update TotalPoints accordingly.
            # Adjust the TotalPoints calculation as needed.
            await conn.execute(
                """
                UPDATE PlayerStats 
                SET ToxicityPoints = ?, TotalPoints = (Participation + Wins - ?)
                WHERE DiscordID = ?
                """,
                (toxicity_points + 1, toxicity_points + 1, discord_id)
            )
            await conn.commit()
            return True  # Successfully updated

        return False  # User not found



# function that removes user from the database
async def remove_user(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # searches for user in the PlayerStats table and saves the result
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()

        # user does not exist in the database
        if result is None:
            return f"This user cannot be found"
        
        await conn.execute("DELETE FROM PlayerStats WHERE DiscordID = ?", (member,))
        await conn.commit()
    
    return f"User {member} has been removed from the database."

# function that unlinks a user's Riot ID from their Discord account
async def unlink_riot_id(discord_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Check if the user exists and has a linked Riot ID
        async with conn.execute(
            "SELECT PlayerRiotID FROM PlayerStats WHERE DiscordID = ?",
            (discord_id,)
        ) as cursor:
            result = await cursor.fetchone()
        
        if result is None:
            return "No account found for this Discord ID."
        
        if result[0] is None:
            return "You don't have a Riot ID linked to your account."
        
        # Store the current Riot ID for the response message
        current_riot_id = result[0]
        
        # Set the PlayerRiotID to NULL
        await conn.execute(
            "UPDATE PlayerStats SET PlayerRiotID = NULL WHERE DiscordID = ?",
            (discord_id,)
        )
        await conn.commit()
        
        return f"Successfully unlinked Riot ID '{current_riot_id}' from your Discord account."

# function that adds an mvp point to a player
async def add_mvp_point(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # tries to find the user in the PlayerStats table and saves the result
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()
        
        # user does not exist in the database
        if result is None:
            return f"This user cannot be found"
        
        await conn.execute("UPDATE PlayerStats SET MVPs = MVPs + 1, TotalPoints = TotalPoints + 1 WHERE DiscordID = ?", (member,))
        await conn.commit()
    
    return f"User {member} has received an MVP point"

# function that sets role preference to the correct Discord user
async def set_role_preference(member, preference: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        # tries to find the user in the PlayerStats table and saves the result
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()
        
        # user does not exist in the database
        if result is None:
            return f"This user cannot be found"

        # adds role preference
        await conn.execute("Update PlayerStats SET RolePreference = ? WHERE DiscordID = ?", (preference, member))
        await conn.commit()

    return f"Updated role preference for {member} to {preference}."

async def update_games_played(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Use member.discord_id instead of member.id for Player objects
        async with conn.execute("SELECT GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(member.discord_id),)) as cursor:
            result = await cursor.fetchone()
        if result:
            games_played = result[0]
            # Update games played
            await conn.execute("UPDATE PlayerStats SET GamesPlayed = ? WHERE DiscordID = ?", (games_played + 1, str(member.discord_id)))
            # Recalculate win rate (wins remains unchanged)
            await update_win_rate(conn, str(member.discord_id))
            await conn.commit()

async def clear_database():
    async with aiosqlite.connect(DB_PATH) as conn:
        await conn.execute("DELETE FROM PlayerStats")
        await conn.commit()
        print("Database cleared successfully.")

class Player:
    def __init__(self, discord_id, username, player_riot_id, participation, wins, mvps,
                 toxicity_points, games_played, win_rate, total_points, tier, rank, role_preference):
        self.discord_id = discord_id
        self.username = username
        self.player_riot_id = player_riot_id
        self.participation = participation
        self.wins = wins
        self.mvps = mvps
        self.toxicity_points = toxicity_points
        self.games_played = games_played
        self.win_rate = win_rate
        self.total_points = total_points
        self.tier = tier
        self.rank = rank
        self.role_preference = role_preference

    def __repr__(self):
        return (f"Player({self.discord_id}, {self.username}, Tier: {self.tier}, Rank: {self.rank}, "
                f"Role Preference: {self.role_preference})")

async def get_player_info(discord_id: str) -> Player:
    """
    Retrieve player information from the database based on DiscordID.
    Returns an instance of the Player class with:
      - discord_id, username, player_riot_id, participation, wins, mvps,
      - toxicity_points, games_played, win_rate, total_points, tier, rank, 
      - role_preference (stored as a list of integers).
    """
    async with aiosqlite.connect(DB_PATH) as conn:
        query = """
        SELECT DiscordID, DiscordUsername, PlayerRiotID, Participation, Wins, MVPs, 
               ToxicityPoints, GamesPlayed, WinRate, TotalPoints, PlayerTier, PlayerRank, RolePreference
        FROM PlayerStats
        WHERE DiscordID = ?
        """
        async with conn.execute(query, (discord_id,)) as cursor:
            result = await cursor.fetchone()

    if result:
        (id_val, username, player_riot_id, participation, wins, mvps, toxicity_points,
         games_played, win_rate, total_points, tier, rank, role_pref_str) = result
        # Convert the role preference string into a list of integers (e.g. "15432" -> [1,5,4,3,2])
        role_preference = [int(ch) for ch in role_pref_str] if role_pref_str else []
        player = Player(
            discord_id=id_val,
            username=username,
            player_riot_id=player_riot_id,
            participation=participation,
            wins=wins,
            mvps=mvps,
            toxicity_points=toxicity_points,
            games_played=games_played,
            win_rate=win_rate,
            total_points=total_points,
            tier=tier,
            rank=rank,
            role_preference=role_preference
        )
        return player
    return None

async def main():
    await initialize_database()
    await clear_database()
    return

if __name__ == "__main__":
    asyncio.run(main())
