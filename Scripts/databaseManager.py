import os
import aiosqlite
from openpyxl import load_workbook
from dotenv import load_dotenv, find_dotenv
import discord
import aiohttp
import asyncio

load_dotenv(find_dotenv())
DB_PATH = os.getenv('DB_PATH')
SPREADSHEET_PATH = os.path.abspath(os.getenv('SPREADSHEET_PATH'))

async def get_db_connection():
    return await aiosqlite.connect(DB_PATH)

async def initialize_database():
    async with aiosqlite.connect(DB_PATH) as conn:
        # Create PlayerStats table
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS "PlayerStats" (
                "DiscordID" TEXT NOT NULL UNIQUE,
                "DiscordUsername" TEXT NOT NULL,
                "PlayerRiotID" TEXT UNIQUE,
                "Participation" NUMERIC DEFAULT 0,
                "Wins" INTEGER DEFAULT 0,
                "MVPs" INTEGER DEFAULT 0,
                "ToxicityPoints" NUMERIC DEFAULT 0,
                "GamesPlayed" INTEGER DEFAULT 0,
                "WinRate" REAL,
                "TotalPoints" NUMERIC DEFAULT 0,
                "PlayerTier" INTEGER DEFAULT 0,
                "PlayerRank" TEXT DEFAULT 'UNRANKED',
                "RolePreference" TEXT DEFAULT NULL,
                PRIMARY KEY("DiscordID")
            )
        ''')
        
        # Create PlayerMatches table to store match history with roles
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS "PlayerMatches" (
                "matchID" INTEGER PRIMARY KEY AUTOINCREMENT,
                "Result" TEXT NOT NULL,
                "blueTop" TEXT NOT NULL,
                "blueJungle" TEXT NOT NULL,
                "blueMid" TEXT NOT NULL,
                "blueBot" TEXT NOT NULL,
                "blueSupport" TEXT NOT NULL,
                "redTop" TEXT NOT NULL,
                "redJungle" TEXT NOT NULL,
                "redMid" TEXT NOT NULL,
                "redBot" TEXT NOT NULL,
                "redSupport" TEXT NOT NULL,
                "MVP" TEXT
            )
        ''')
        
        await conn.commit()

def update_excel(discord_id, player_data):
    try:
        # Load the workbook and get the correct sheet
        workbook = load_workbook(SPREADSHEET_PATH)
        sheet_name = 'PlayerStats'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f'Sheet {sheet_name} does not exist in the workbook')

        # Check if the player already exists in the sheet
        found = False
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
            if str(row[0].value) == discord_id:  # Check if Discord ID matches
                # Update only if there's a difference
                for idx, key in enumerate(player_data.keys()):
                    if row[idx].value != player_data[key]:
                        row[idx].value = player_data[key]
                        found = True
                break

        # If player not found, add a new row
        if not found:
            # Find the first truly empty row, ignoring formatting
            empty_row_idx = sheet.max_row + 1
            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if all(cell.value is None for cell in row):
                    empty_row_idx = i
                    break

            # Insert the new data into the empty row
            for idx, key in enumerate(player_data.keys(), start=1):
                sheet.cell(row=empty_row_idx, column=idx).value = player_data[key]

        # Save the workbook after updates
        workbook.save(SPREADSHEET_PATH)
        print(f"Spreadsheet '{SPREADSHEET_PATH}' has been updated successfully.")

    except Exception as e:
        print(f"Error updating Excel file: {e}")

async def check_winners_in_db(winners):
    not_found_users = []
    async with aiosqlite.connect(DB_PATH) as conn:
        for winner in winners:
            # Use winner.discord_id instead of winner.id
            async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(winner.discord_id),)) as cursor:
                result = await cursor.fetchone()

            if not result:
                # Add to not found list
                not_found_users.append(winner)


async def update_wins(winners):
    async with aiosqlite.connect(DB_PATH) as conn:
        for winner in winners:
            # Use winner.discord_id instead of winner.id
            async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(winner.discord_id),)) as cursor:
                result = await cursor.fetchone()
                if result:
                    wins, games_played = result
                    await conn.execute(
                        "UPDATE PlayerStats SET Wins = ?, GamesPlayed = ? WHERE DiscordID = ?",
                        (wins + 1, games_played + 1, str(winner.discord_id))
                    )
                    await update_win_rate(conn, str(winner.discord_id))
        await conn.commit()



# code to calculate and update winrate in database
async def update_win_rate(conn, discord_id):
    async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (discord_id,)) as cursor:
        result = await cursor.fetchone()
    if result:
        wins, games_played = result
        win_rate = (wins / games_played) * 100 if games_played > 0 else 0
        await conn.execute("UPDATE PlayerStats SET WinRate = ? WHERE DiscordID = ?", (win_rate, discord_id))

# discord roles need to be implemented here
# maybe adding functionality for multiple users 
async def update_points(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()
        
        if result is None:
            return f"Failed to update: {member} (user not found in database)."
        
        await conn.execute("UPDATE PlayerStats SET Participation = Participation + 1, TotalPoints = TotalPoints + 1 WHERE DiscordID = ?", (member,))
        await conn.commit()
        return f"Added 1 participation point to {member}."

async def check_and_update_rank(player_id: str, riot_id: str):
    """
    Check if a player's rank has changed by fetching current rank from Riot API.
    Returns a tuple (bool, str) where the first value indicates if the rank changed,
    and the second value is a message about the change.
    """
    # Get API key
    api_key = os.getenv("RIOT_API_KEY")
    if not api_key:
        return False, "Could not check rank: API key not found."
    
    # Extract summoner name and tagline from riot_id
    if '#' not in riot_id:
        return False, "Invalid Riot ID format."
    summoner_name, tagline = riot_id.split('#', 1)
    
    try:
        # Get current rank in database
        async with aiosqlite.connect(DB_PATH) as conn:
            async with conn.execute(
                "SELECT PlayerRank FROM PlayerStats WHERE DiscordID = ?",
                (player_id,)
            ) as cursor:
                result = await cursor.fetchone()
            
            if not result:
                return False, "Player not found in database."
            
            current_db_rank = result[0]
            
            # Fetch current rank from Riot API
            headers = {"X-Riot-Token": api_key}
            
            # Step 1: Get account info
            async with aiohttp.ClientSession() as session:
                account_url = f"https://americas.api.riotgames.com/riot/account/v1/accounts/by-riot-id/{summoner_name}/{tagline}"
                async with session.get(account_url, headers=headers) as response:
                    if response.status != 200:
                        return False, "Could not verify Riot ID."
                    
                    account_data = await response.json()
                    puuid = account_data.get("puuid")
                    
                    # Step 2: Get Summoner ID
                    summoner_url = f"https://na1.api.riotgames.com/lol/summoner/v4/summoners/by-puuid/{puuid}"
                    async with session.get(summoner_url, headers=headers) as summoner_response:
                        if summoner_response.status != 200:
                            return False, "Could not fetch summoner data."
                        
                        summoner_data = await summoner_response.json()
                        summoner_id = summoner_data.get("id")
                        
                        # Step 3: Get ranked stats
                        ranked_url = f"https://na1.api.riotgames.com/lol/league/v4/entries/by-summoner/{summoner_id}"
                        async with session.get(ranked_url, headers=headers) as ranked_response:
                            if ranked_response.status != 200:
                                return False, "Could not fetch ranked data."
                            
                            ranked_data = await ranked_response.json()
                            
                            # Look for SOLO_DUO queue data
                            api_rank = "UNRANKED"
                            for queue_data in ranked_data:
                                if queue_data.get("queueType") == "RANKED_SOLO_5x5":
                                    api_rank = queue_data.get("tier", "UNRANKED")
                                    break
                            
                            # Map rank to tier
                            rank_to_tier = {
                                "IRON": 7,
                                "BRONZE": 6,
                                "SILVER": 6,
                                "GOLD": 5,
                                "PLATINUM": 4,
                                "EMERALD": 3,
                                "DIAMOND": 3,
                                "MASTER": 2,
                                "GRANDMASTER": 1,
                                "CHALLENGER": 1
                            }
                            api_tier = rank_to_tier.get(api_rank, 7)
                            
                            # Check if rank changed
                            if current_db_rank != api_rank:
                                # Update database with new rank
                                await conn.execute(
                                    "UPDATE PlayerStats SET PlayerRank = ?, PlayerTier = ? WHERE DiscordID = ?",
                                    (api_rank, api_tier, player_id)
                                )
                                await conn.commit()
                                
                                return True, f"Your rank has changed from {current_db_rank} to {api_rank}!"
                            
                            return False, "Your rank has not changed."
                        
    except Exception as e:
        print(f"Error checking rank: {e}")
        return False, f"Error checking rank: {str(e)}"

async def update_username(player):
    try:
        current_username = player.display_name
        async with aiosqlite.connect(DB_PATH) as conn:
            # Fetch the player's current data from the database
            async with conn.execute(
                "SELECT DiscordUsername FROM PlayerStats WHERE DiscordID = ?",
                (str(player.id),)
            ) as cursor:
                player_stats = await cursor.fetchone()

            # If player exists in the database and the username is outdated, update it
            if player_stats:
                stored_username = player_stats[0]
                if stored_username != current_username:
                    await conn.execute(
                        "UPDATE PlayerStats SET DiscordUsername = ? WHERE DiscordID = ?",
                        (current_username, str(player.id))
                    )
                    await conn.commit()
                    print(f"Updated username for {player.id} to '{current_username}'.")
            else:
                # Insert new player if they don't exist
                await conn.execute(
                    """INSERT INTO PlayerStats (DiscordID, DiscordUsername)
                       VALUES (?, ?)""",
                    (str(player.id), current_username)
                )
                await conn.commit()
                print(f"Added new player {player.id} with username '{current_username}'.")
    except Exception as e:
        print(f"Error updating username: {e}")

# member was interaction: discord.Interaction
async def link(member, riot_id: str):
    """
    Link a Discord user's account with their Riot ID.
    
    Args:
        member: The Discord member object
        riot_id: The Riot ID in format 'username#tagline'
    
    Returns:
        A message string indicating the result of the operation
    """
    # Riot ID is in the format 'username#tagline', e.g., 'jacstogs#1234'
    if '#' not in riot_id:
        return "Invalid Riot ID format. Must include a '#' character (e.g., 'username#tagline')."

    # Split the Riot ID into name and tagline
    summoner_name, tagline = riot_id.split('#', 1)
    summoner_name = summoner_name.strip()
    tagline = tagline.strip()
    
    # Check if either part is empty
    if not summoner_name or not tagline:
        return "Invalid Riot ID format. Both username and tagline must be provided."

    # Check if API key exists
    api_key = os.getenv("RIOT_API_KEY")
    if not api_key:
        return "ERROR: Riot API key not found. Please add a valid API key to your environment variables."
    
    # Verify that the Riot ID exists using the Riot API
    headers = {"X-Riot-Token": api_key}
    account_url = f"https://americas.api.riotgames.com/riot/account/v1/accounts/by-riot-id/{summoner_name}/{tagline}"
    verified = False  # Default to not verified
    player_rank = "UNRANKED"
    player_tier = 7  # Default to lowest tier

    try:
        async with aiohttp.ClientSession() as session:
            # Step 1: Get account info
            async with session.get(account_url, headers=headers) as response:
                if response.status == 200:
                    # Riot ID exists
                    account_data = await response.json()
                    verified = True
                    puuid = account_data.get("puuid")
                    
                    # Step 2: Get Summoner ID from PUUID (for NA1 region)
                    summoner_url = f"https://na1.api.riotgames.com/lol/summoner/v4/summoners/by-puuid/{puuid}"
                    async with session.get(summoner_url, headers=headers) as summoner_response:
                        if summoner_response.status == 200:
                            summoner_data = await summoner_response.json()
                            summoner_id = summoner_data.get("id")
                            
                            # Step 3: Get ranked stats (focusing on SOLO_DUO queue)
                            ranked_url = f"https://na1.api.riotgames.com/lol/league/v4/entries/by-summoner/{summoner_id}"
                            async with session.get(ranked_url, headers=headers) as ranked_response:
                                if ranked_response.status == 200:
                                    ranked_data = await ranked_response.json()
                                    
                                    # Look for SOLO_DUO queue data
                                    for queue_data in ranked_data:
                                        if queue_data.get("queueType") == "RANKED_SOLO_5x5":
                                            player_rank = queue_data.get("tier", "UNRANKED")
                                            
                                            # Map rank to tier based on matchmaking logic
                                            rank_to_tier = {
                                                "IRON": 7,
                                                "BRONZE": 6,
                                                "SILVER": 6,
                                                "GOLD": 5,
                                                "PLATINUM": 4,
                                                "EMERALD": 3,
                                                "DIAMOND": 3,
                                                "MASTER": 2,
                                                "GRANDMASTER": 1,
                                                "CHALLENGER": 1
                                            }
                                            player_tier = rank_to_tier.get(player_rank, 7)
                                            break
                                else:
                                    print(f"Warning: Could not fetch ranked data. Status: {ranked_response.status}")
                        else:
                            print(f"Warning: Could not fetch summoner data. Status: {summoner_response.status}")
                else:
                    # Failed to verify Riot ID
                    error_msg = await response.text()
                    if response.status == 404:
                        return f"The Riot ID '{riot_id}' could not be found. Please double-check and try again."
                    elif response.status == 400:
                        return f"ERROR: Bad Request. This is likely due to an invalid or expired API key format."
                    elif response.status == 401:
                        return f"ERROR: Invalid Riot API key. Please verify your API key or obtain a new one from the Riot Developer Portal."
                    elif response.status == 403:
                        return f"ERROR: Riot API key has expired or doesn't have sufficient permissions. Please obtain a new key from the Riot Developer Portal."
                    else:
                        return f"ERROR: API Error (status {response.status}). This may be due to API key issues or service problems."
    except aiohttp.ClientError as e:
        return f"Network error connecting to Riot API: {e}. Please check your internet connection and try again."
    except Exception as e:
        return f"Unexpected error during API validation: {e}. Please try again later."

    # Only proceed if verification passed
    if verified:
        async with aiosqlite.connect(DB_PATH) as conn:
            try:
                # Use member's id for database operations
                async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
                    result = await cursor.fetchone()

                if result:
                    # Update the existing record with the new Riot ID and rank information
                    await conn.execute(
                        """
                        UPDATE PlayerStats
                        SET PlayerRiotID = ?, PlayerRank = ?, PlayerTier = ?
                        WHERE DiscordID = ?
                        """,
                        (riot_id, player_rank, player_tier, str(member.id))
                    )
                else:
                    # Insert a new record if the user doesn't exist in the database
                    await conn.execute(
                        """
                        INSERT INTO PlayerStats
                        (DiscordID, DiscordUsername, PlayerRiotID, PlayerRank, PlayerTier)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (str(member.id), member.display_name, riot_id, player_rank, player_tier)
                    )

                await conn.commit()
                
                # Formulate a message including the rank
                rank_message = f" Your rank has been determined to be {player_rank}."
                if player_rank == "UNRANKED":
                    rank_message = " Your rank could not be determined or you're unranked."
                    
                return f"Your Riot ID '{riot_id}' has been successfully linked to your Discord account.{rank_message}"

            except aiosqlite.IntegrityError as e:
                # Handle UNIQUE constraint violation (i.e., Riot ID already linked)
                if 'UNIQUE constraint failed: PlayerStats.PlayerRiotID' in str(e):
                    # Riot ID is already linked to another user
                    async with conn.execute("""
                        SELECT DiscordID, DiscordUsername FROM PlayerStats WHERE PlayerRiotID = ?
                    """, (riot_id,)) as cursor:
                        existing_user_data = await cursor.fetchone()

                    if existing_user_data:
                        existing_user_id, existing_username = existing_user_data
                        return f"Error: This Riot ID is already linked to another Discord user: <@{existing_user_id}>. If this is a mistake, please contact an administrator."
                else:
                    return f"Database error: {e}"
            except Exception as e:
                return f"An unexpected error occurred while updating the database: {e}"
    # This else clause is not needed since all code paths in the if block return a value

async def update_toxicity_by_id(discord_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Search for the user in the PlayerStats table using the given discord_id
        async with conn.execute("SELECT ToxicityPoints FROM PlayerStats WHERE DiscordID = ?", (discord_id,)) as cursor:
            result = await cursor.fetchone()

        if result:
            toxicity_points = result[0]
            # Increment the toxicity points and update TotalPoints accordingly.
            # Adjust the TotalPoints calculation as needed.
            await conn.execute(
                """
                UPDATE PlayerStats 
                SET ToxicityPoints = ?, TotalPoints = (Participation + Wins - ?)
                WHERE DiscordID = ?
                """,
                (toxicity_points + 1, toxicity_points + 1, discord_id)
            )
            await conn.commit()
            return True  # Successfully updated

        return False  # User not found



# function that removes user from the database
async def remove_user(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # searches for user in the PlayerStats table and saves the result
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()

        # user does not exist in the database
        if result is None:
            return f"This user cannot be found"
        
        await conn.execute("DELETE FROM PlayerStats WHERE DiscordID = ?", (member,))
        await conn.commit()
    
    return f"User {member} has been removed from the database."

# function that unlinks a user's Riot ID from their Discord account
async def unlink_riot_id(discord_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Check if the user exists and has a linked Riot ID
        async with conn.execute(
            "SELECT PlayerRiotID FROM PlayerStats WHERE DiscordID = ?",
            (discord_id,)
        ) as cursor:
            result = await cursor.fetchone()
        
        if result is None:
            return "No account found for this Discord ID."
        
        if result[0] is None:
            return "You don't have a Riot ID linked to your account."
        
        # Store the current Riot ID for the response message
        current_riot_id = result[0]
        
        # Set the PlayerRiotID to NULL
        await conn.execute(
            "UPDATE PlayerStats SET PlayerRiotID = NULL WHERE DiscordID = ?",
            (discord_id,)
        )
        await conn.commit()
        
        return f"Successfully unlinked Riot ID '{current_riot_id}' from your Discord account."

# Store match data with player roles in the PlayerMatches table
async def store_match_data(game_data, result, mvp_id=None):
    """
    Store match data in the PlayerMatches table with role assignments.
    
    Args:
        game_data: Dictionary containing 'blue' and 'red' team player lists
        result: String indicating which team won ('blue' or 'red')
        mvp_id: Discord ID of the MVP player (or None if MVP voting was skipped)
    """
    async with aiosqlite.connect(DB_PATH) as conn:
        # Player arrays are already in role order: Top, Jungle, Mid, Bot, Support
        blue_team = game_data["blue"]
        red_team = game_data["red"]
        
        # Ensure both teams have 5 players
        if len(blue_team) != 5 or len(red_team) != 5:
            print(f"Error: Teams must have exactly 5 players. Blue: {len(blue_team)}, Red: {len(red_team)}")
            return False
        
        # Insert match data with role assignments
        await conn.execute("""
            INSERT INTO PlayerMatches (
                Result,
                blueTop, blueJungle, blueMid, blueBot, blueSupport,
                redTop, redJungle, redMid, redBot, redSupport,
                MVP
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            result,
            blue_team[0].discord_id, blue_team[1].discord_id, blue_team[2].discord_id,
            blue_team[3].discord_id, blue_team[4].discord_id,
            red_team[0].discord_id, red_team[1].discord_id, red_team[2].discord_id,
            red_team[3].discord_id, red_team[4].discord_id,
            mvp_id
        ))
        
        await conn.commit()
        return True

# Comprehensive function to update all player stats after a match
async def update_all_player_stats(game_data, result, mvp_id=None):
    """
    Update all player statistics after a match completes.
    
    Args:
        game_data: Dictionary containing 'blue' and 'red' team player lists
        result: String indicating which team won ('blue' or 'red')
        mvp_id: Discord ID of the MVP player (or None if MVP voting was skipped)
    """
    async with aiosqlite.connect(DB_PATH) as conn:
        blue_team = game_data["blue"]
        red_team = game_data["red"]
        all_players = blue_team + red_team
        
        # Determine winning team
        winners = blue_team if result == "blue" else red_team
        
        # Update all players' participation and games played
        for player in all_players:
            player_id = str(player.discord_id)
            
            # Check if player exists in database
            async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (player_id,)) as cursor:
                player_exists = await cursor.fetchone()
            
            if not player_exists:
                print(f"Warning: Player {player_id} not found in database.")
                continue
            
            # Update participation for all players
            await conn.execute(
                "UPDATE PlayerStats SET Participation = Participation + 1 WHERE DiscordID = ?",
                (player_id,)
            )
            
            # Update games played for all players
            await conn.execute(
                "UPDATE PlayerStats SET GamesPlayed = GamesPlayed + 1 WHERE DiscordID = ?",
                (player_id,)
            )
        
        # Update wins for winning team
        for player in winners:
            player_id = str(player.discord_id)
            await conn.execute(
                "UPDATE PlayerStats SET Wins = Wins + 1 WHERE DiscordID = ?",
                (player_id,)
            )
        
        # Update MVP stat if applicable
        if mvp_id:
            await conn.execute(
                "UPDATE PlayerStats SET MVPs = MVPs + 1, TotalPoints = TotalPoints + 1 WHERE DiscordID = ?",
                (mvp_id,)
            )
        
        # Update TotalPoints for all players
        for player in all_players:
            player_id = str(player.discord_id)
            
            # TotalPoints = Participation + Wins + MVPs - ToxicityPoints
            await conn.execute("""
                UPDATE PlayerStats
                SET TotalPoints = Participation + Wins + MVPs - ToxicityPoints
                WHERE DiscordID = ?
            """, (player_id,))
            
            # Update win rate
            await update_win_rate(conn, player_id)
        
        await conn.commit()

# function that adds an mvp point to a player
# This function is kept for backward compatibility but is no longer used directly
# MVP points are now handled by update_all_player_stats
async def add_mvp_point(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # tries to find the user in the PlayerStats table and saves the result
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()
        
        # user does not exist in the database
        if result is None:
            return f"This user cannot be found"
        
        await conn.execute("UPDATE PlayerStats SET MVPs = MVPs + 1, TotalPoints = TotalPoints + 1 WHERE DiscordID = ?", (member,))
        await conn.commit()
    
    return f"User {member} has received an MVP point"

# function that sets role preference to the correct Discord user
async def set_role_preference(member, preference: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        # tries to find the user in the PlayerStats table and saves the result
        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (member,)) as cursor:
            result = await cursor.fetchone()
        
        # user does not exist in the database
        if result is None:
            return f"This user cannot be found"

        # adds role preference
        await conn.execute("Update PlayerStats SET RolePreference = ? WHERE DiscordID = ?", (preference, member))
        await conn.commit()

    return f"Updated role preference for {member} to {preference}."

async def update_games_played(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Use member.discord_id instead of member.id for Player objects
        async with conn.execute("SELECT GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(member.discord_id),)) as cursor:
            result = await cursor.fetchone()
        if result:
            games_played = result[0]
            # Update games played
            await conn.execute("UPDATE PlayerStats SET GamesPlayed = ? WHERE DiscordID = ?", (games_played + 1, str(member.discord_id)))
            # Recalculate win rate (wins remains unchanged)
            await update_win_rate(conn, str(member.discord_id))
            await conn.commit()

async def clear_database():
    async with aiosqlite.connect(DB_PATH) as conn:
        await conn.execute("DELETE FROM PlayerStats")
        await conn.commit()
        print("Database cleared successfully.")

async def add_30_players_with_ranks():
    # Ranks from lowest to highest
    ranks = ["IRON", "BRONZE", "SILVER", "GOLD", "PLATINUM", "EMERALD", "DIAMOND", "MASTER", "GRANDMASTER", "CHALLENGER"]
    roles = ["Top", "Jungle", "Mid", "Bot", "Support"]
    # Mapping of most desirable role to a sample 5-character preference string
    role_to_pref = {
        "Top": "15432",     # '1' in position 0 (Top is most desired)
        "Jungle": "51432",  # '1' in position 1 (Jungle is most desired)
        "Mid": "54132",     # '1' in position 2 (Mid is most desired)
        "Bot": "54312",     # '1' in position 3 (Bot is most desired)
        "Support": "54321"  # '1' in position 4 (Support is most desired)
    }
    
    # Map ranks to tiers (lower tier number = higher skill)
    # Based on Matchmaking.py logic
    rank_to_tier = {
        "IRON": 7,
        "BRONZE": 6,
        "SILVER": 6,
        "GOLD": 5,
        "PLATINUM": 4,
        "EMERALD": 3,
        "DIAMOND": 3,
        "MASTER": 2,
        "GRANDMASTER": 1,
        "CHALLENGER": 1
    }
    
    async with aiosqlite.connect(DB_PATH) as conn:
        # Clear existing players first
        await conn.execute("DELETE FROM PlayerStats")
        
        # Distribute 30 players across all ranks
        for i in range(1, 31):  # IDs 1 to 30
            # Determine rank (distribute players across all ranks)
            rank_index = min(9, (i-1) // 3)  # 3 players per rank
            rank = ranks[rank_index]
            
            # Get corresponding tier
            tier = rank_to_tier[rank]
            
            # Cycle through the roles for most desirable assignment
            most_desirable = roles[(i-1) % 5]
            
            # Construct username: id + rank + role (more concise format)
            username = f"P{i}_{rank[0:3]}_{most_desirable[0]}"
            
            # Set role preference
            role_pref = role_to_pref[most_desirable]
            
            # Create a fake Riot ID
            riot_id = f"player{i}#{1000+i}"
            
            # Insert the player into the PlayerStats table with appropriate rank and tier
            await conn.execute(
                """
                INSERT INTO PlayerStats (
                    DiscordID, DiscordUsername, PlayerRiotID,
                    PlayerTier, PlayerRank, RolePreference
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (str(i), username, riot_id, tier, rank, role_pref)
            )
        
        await conn.commit()
        print("30 players with ranks have been added to the database.")

class Player:
    def __init__(self, discord_id, username, player_riot_id, participation, wins, mvps,
                 toxicity_points, games_played, win_rate, total_points, tier, rank, role_preference):
        self.discord_id = discord_id
        self.username = username
        self.player_riot_id = player_riot_id
        self.participation = participation
        self.wins = wins
        self.mvps = mvps
        self.toxicity_points = toxicity_points
        self.games_played = games_played
        self.win_rate = win_rate
        self.total_points = total_points
        self.tier = tier
        self.rank = rank
        self.role_preference = role_preference

    def __repr__(self):
        return (f"Player({self.discord_id}, {self.username}, Tier: {self.tier}, Rank: {self.rank}, "
                f"Role Preference: {self.role_preference})")

async def get_player_info(discord_id: str) -> Player:
    """
    Retrieve player information from the database based on DiscordID.
    Returns an instance of the Player class with:
      - discord_id, username, player_riot_id, participation, wins, mvps,
      - toxicity_points, games_played, win_rate, total_points, tier, rank, 
      - role_preference (stored as a list of integers).
    """
    async with aiosqlite.connect(DB_PATH) as conn:
        query = """
        SELECT DiscordID, DiscordUsername, PlayerRiotID, Participation, Wins, MVPs, 
               ToxicityPoints, GamesPlayed, WinRate, TotalPoints, PlayerTier, PlayerRank, RolePreference
        FROM PlayerStats
        WHERE DiscordID = ?
        """
        async with conn.execute(query, (discord_id,)) as cursor:
            result = await cursor.fetchone()

    if result:
        (id_val, username, player_riot_id, participation, wins, mvps, toxicity_points,
         games_played, win_rate, total_points, tier, rank, role_pref_str) = result
        # Convert the role preference string into a list of integers (e.g. "15432" -> [1,5,4,3,2])
        role_preference = [int(ch) for ch in role_pref_str] if role_pref_str else []
        player = Player(
            discord_id=id_val,
            username=username,
            player_riot_id=player_riot_id,
            participation=participation,
            wins=wins,
            mvps=mvps,
            toxicity_points=toxicity_points,
            games_played=games_played,
            win_rate=win_rate,
            total_points=total_points,
            tier=tier,
            rank=rank,
            role_preference=role_preference
        )
        return player
    return None

async def main():
    await initialize_database()
    await clear_database()
    await add_30_players_with_ranks()
    return

if __name__ == "__main__":
    asyncio.run(main())
